import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
from sklearn.preprocessing import LabelEncoder, StandardScaler
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import threading
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import mplcursors
import matplotlib as mpl


pd.set_option('future.no_silent_downcasting', True) #Включает строгий режим обработки неявного приведения типов в pandas
plt.rcParams.update({
    'font.size': 9,
    'axes.formatter.useoffset': False,
    'axes.formatter.limits': (-4, 4)
})

def sanitize_filename(filename): #Очищает строку от недопустимых символов для использования в имени файла
    return re.sub(r'[<>:"/\\|?*]', '_', filename)

class SimpleNN: #Простая нейросеть с ReLU-активацией для регрессии спроса
    def __init__(self, input_dim, lr=0.02, epochs=100, batch_size=32): #Инициализация параметров нейросети
        self.W = np.random.randn(input_dim, 1) * np.sqrt(2.0 / input_dim) #Инициализация весов по методу Хе, оптимально для ReLU
        self.b = np.zeros((1,))
        self.lr = lr
        self.epochs = epochs
        self.batch_size = batch_size
        self.input_dim = input_dim

    def relu(self, z): #Функция активации ReLU
        return np.maximum(0, z)

    def relu_derivative(self, z): #Производная ReLU: 1 если z > 0, иначе 0
        return (z > 0).astype(np.float64)

    def forward(self, X): #Прямой проход: вычисление прогноза по входным данным
        if X.shape[1] != self.input_dim:
            raise ValueError(f"Ожидалось {self.input_dim} признаков, получено {X.shape[1]}")
        z = X @ self.W + self.b #Линейное преобразование
        a = self.relu(z) #Применение нелинейности
        return a, z

    def fit(self, X, y, residuals=None): #Обучение модели градиентным бустингом
        targets = residuals if residuals is not None else y.reshape(-1, 1) #Определяем целевые значения: либо остатки, либо исходный y
        n = X.shape[0]
        if self.W.shape[1] != 1: #Гарантируем корректную форму весов и смещения
            self.W = self.W[:, :1].copy()
        if self.b.ndim == 0:
            self.b = np.array([self.b])
        for epoch in range(self.epochs): #Обучение в течение заданного числа эпох
            idx = np.random.permutation(n) #Перемешиваем данные для стабильности обучения
            X_shuf, y_shuf = X[idx], targets[idx]
            for i in range(0, n, self.batch_size): #Обучение по мини-батчам
                X_batch = X_shuf[i:i + self.batch_size]
                y_batch = y_shuf[i:i + self.batch_size]
                y_pred, z_batch = self.forward(X_batch) #Прямой проход
                error = y_batch - y_pred #Вычисление ошибки
                dz = error * self.relu_derivative(z_batch) #Обратное распространение: градиент по z с учётом производной ReLU
                dW = -X_batch.T @ dz / X_batch.shape[0] #Градиенты по весам и смещению
                db = -np.mean(dz, axis=0, keepdims=True)
                if dW.shape[1] != 1: #Коррекция формы градиентов
                    dW = dW[:, :1]
                if db.shape != (1, 1):
                    db = db[:1, :1]
                self.W -= self.lr * dW  #Обновление параметров модели
                self.b -= self.lr * db[0, 0]

    def predict(self, X): #Получение прогноза для новых данных
        y_pred, _ = self.forward(X)
        return y_pred.flatten() #Преобразуем в 1D-массив

class NNBoost: #Реализация градиентного бустинга на основе нейронных сетей
    def __init__(self, n_estimators=50, learning_rate=0.03): #Инициализация параметров бустинга
        self.models = []
        self.n_estimators = n_estimators
        self.lr = learning_rate

    def fit(self, X, y): #Обучение ансамбля моделей методом градиентного бустинга
        residuals = y.astype(np.float64).copy() #Копируем целевую переменную как начальные "остатки"
        for t in range(self.n_estimators):
            model = SimpleNN(X.shape[1], lr=0.02, epochs=100) #Создаём новую нейросеть для текущей итерации
            model.fit(X, y, residuals=residuals) #Обучаем модель на остатках
            pred = model.predict(X) #Получаем прогноз текущей модели
            residuals -= self.lr * pred #Обновляем остатки: вычитаем скорректированный прогноз
            self.models.append(model) #Сохраняем модель в ансамбль

    def predict(self, X): #Формирование итогового прогноза путём суммирования прогнозов всех моделей.
        pred = np.zeros(X.shape[0]) #Инициализируем нулевой прогноз
        for model in self.models: #Суммируем прогнозы всех моделей с учётом learning_rate
            pred += self.lr * model.predict(X)
        return pred

class SalesAnalyzerApp: #Основной класс GUI-приложения для анализа продаж, прогнозирования спроса и выручки
    def __init__(self, root): #Инициализация главного окна и всех UI-компонентов
        self.root = root
        self.root.title("Анализатор продаж")
        self.root.resizable(False, False) #Фиксированный размер окна
        self.root.geometry("1630x715")
        self.root.configure(bg="#f8f9fa")
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing) #Обработка закрытия
        self.top_frame = tk.Frame(root, bg="#f8f9fa", pady=5)
        self.top_frame.pack(fill=tk.X)
        self.status_label = tk.Label(
            self.top_frame,
            text="Прогноз на 2026 год (12 мес.)",
            font=("Arial", 14, "bold"),
            fg="#1a5276",
            bg="#f8f9fa"
        )
        self.status_label.pack(side=tk.LEFT, padx=20)
        self.theme_btn = tk.Button( #Кнопка переключения темы (светлая/тёмная)
            self.top_frame,
            text="Тёмная тема",
            command=self.toggle_theme,
            bg="#7f8c8d",
            fg="white",
            width=10,
            font=("Arial", 12)
        )
        self.theme_btn.pack(side=tk.RIGHT, padx=10)
        self.dark_theme = False
        self.control_frame = tk.Frame(root, bg="#f8f9fa", pady=5) #Панель управления: загрузка файла и запуск анализа
        self.control_frame.pack(fill=tk.X)
        self.center_container = tk.Frame(self.control_frame, bg="#f8f9fa")
        self.center_container.pack(expand=True)
        tk.Label(self.center_container, text="Файл:", font=("Arial", 11), bg="#f8f9fa").pack(side=tk.LEFT, padx=5) #Поле ввода пути к файлу
        self.path_entry = tk.Entry(self.center_container, width=65, font=("Consolas", 10))
        self.path_entry.pack(side=tk.LEFT, padx=5)
        self.path_entry.insert(0, "")
        self.browse_btn = tk.Button( #Кнопка выбора файла
            self.center_container, text="Выбрать", command=self.browse_file,
            bg="#3498db", fg="white", padx=10
        )
        self.browse_btn.pack(side=tk.LEFT, padx=5)
        self.load_btn = tk.Button( #Кнопка запуска анализа
            self.center_container, text="Запустить анализ",
            command=self.start_analysis,
            bg="#27ae60", fg="white", font=("Arial", 10, "bold"), padx=15
        )
        self.load_btn.pack(side=tk.LEFT, padx=5)
        self.export_btn = tk.Button( #Кнопки экспорта (изначально неактивны)
            self.center_container, text="Сохранить в Excel",
            command=self.export_to_excel,
            bg="#218359", fg="white", font=("Arial", 10, "bold"), padx=10
        )
        self.export_btn.pack(side=tk.LEFT, padx=5)
        self.export_btn.config(state="disabled") #Активируется после анализа
        self.export_plot_btn = tk.Button(
            self.center_container, text="Сохранить графики",
            command=self.export_plots,
            bg="#21837a", fg="white", font=("Arial", 10, "bold"), padx=10
        )
        self.export_plot_btn.pack(side=tk.LEFT, padx=5)
        self.export_plot_btn.config(state="disabled")
        self.seasonality_btn = tk.Button( #Кнопки дополнительных отчётов (изначально неактивны)
            self.center_container, text="Сезонность",
            command=self.show_seasonality,
            bg="#2f7f99", fg="white", font=("Arial", 10, "bold"), padx=10
        )
        self.seasonality_btn.pack(side=tk.LEFT, padx=5)
        self.seasonality_btn.config(state="disabled")
        self.elasticity_btn = tk.Button(
            self.center_container, text="Эластичность",
            command=self.show_elasticity,
            bg="#127d87", fg="white", font=("Arial", 10, "bold"), padx=10
        )
        self.elasticity_btn.pack(side=tk.LEFT, padx=5)
        self.elasticity_btn.config(state="disabled")
        cat_frame = tk.Frame(self.control_frame, bg="#f8f9fa", pady=5) #Выбор категорий товаров
        cat_frame.pack(expand=True)
        tk.Label(cat_frame, text="Категории:", font=("Arial", 11), bg="#f8f9fa").pack(side=tk.LEFT, padx=5)
        list_frame = tk.Frame(cat_frame) #Список категорий с прокруткой
        list_frame.pack(side=tk.LEFT, padx=5)
        self.cat_listbox = tk.Listbox(
            list_frame,
            selectmode=tk.EXTENDED, #Множественный выбор
            height=5,
            width=28,
            font=("Arial", 10),
            exportselection=False #Сохраняет выделение при фокусе вне списка
        )
        scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.cat_listbox.yview)
        self.cat_listbox.config(yscrollcommand=scrollbar.set)
        self.cat_listbox.pack(side=tk.LEFT, fill=tk.BOTH)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.apply_cat_btn = tk.Button( #Кнопка применения выбора категорий
            cat_frame, text="Применить", command=self.apply_category_selection,
            bg="#3498db", fg="white", padx=10, font=("Arial", 9)
        )
        self.apply_cat_btn.pack(side=tk.LEFT, padx=(10, 0))
        self.cat_listbox.insert(tk.END, "Все категории") #По умолчанию выбрана опция "Все категории"
        self.cat_listbox.select_set(0)
        self.selected_categories = ["Все категории"]
        self.df_full_raw = None #Хранение данных и состояния
        self.df_full = None
        self.monthly_df = None
        self.future_monthly = None
        self.monthly_by_cat = {}
        self.future_by_cat = {}
        self.is_loading = False
        self.anim1 = self.anim2 = None
        self.current_view = "charts"
        self.anim3_bars = []
        self.anim_elasticity_scatters = []
        self.seasonal_coeffs = {}
        self.charts_frame = tk.Frame(root, bg="#f8f9fa") #Основная область: графики
        self.charts_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.stats_frame = tk.Frame(root, bg="#f8f9fa") #Строка статистики под графиками
        self.stats_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        self.stats_label = tk.Label(
            self.stats_frame,
            text="",
            font=("Arial", 11),
            fg="#2c3e50",
            bg="#f8f9fa",
            justify=tk.LEFT
        )
        self.stats_label.pack(side=tk.LEFT)
        self.seasonality_frame = tk.Frame(root, bg="#f8f9fa") #Вкладка "Сезонность"
        self.seasonality_top_frame = tk.Frame(self.seasonality_frame, bg="#f8f9fa") #Верхняя панель вкладки сезонности
        self.seasonality_top_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
        self.back_from_seasonality_btn = tk.Button(
            self.seasonality_top_frame,
            text="Назад к графикам",
            command=self.show_charts,
            bg="#2c3e50", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5
        )
        self.back_from_seasonality_btn.pack(side=tk.RIGHT, padx=(5, 0))
        seasonality_title = tk.Label(
            self.seasonality_top_frame,
            text="Сезонность спроса по месяцам",
            font=("Arial", 14, "bold"),
            fg="#1a5276",
            bg="#f8f9fa"
        )
        seasonality_title.pack(side=tk.LEFT, padx=10)
        self.fig1, self.ax1 = plt.subplots(figsize=(6.5, 4.8))
        self.canvas1 = FigureCanvasTkAgg(self.fig1, self.charts_frame)
        self.canvas1.get_tk_widget().grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        self.fig2, self.ax2 = plt.subplots(figsize=(6.5, 4.8))
        self.canvas2 = FigureCanvasTkAgg(self.fig2, self.charts_frame)
        self.canvas2.get_tk_widget().grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        self.charts_frame.grid_columnconfigure(0, weight=1)
        self.charts_frame.grid_columnconfigure(1, weight=1)
        self.seasonality_content_frame = tk.Frame(self.seasonality_frame, bg="#f8f9fa") #График и текстовая панель сезонности
        self.seasonality_content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.seasonality_graph_frame = tk.Frame(self.seasonality_content_frame, bg="#f8f9fa")
        self.seasonality_graph_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        self.seasonality_text_frame = tk.Frame(self.seasonality_content_frame, bg="#f8f9fa", width=750, height=340) #Текстовая панель рекомендаций
        self.seasonality_text_frame.pack(side=tk.RIGHT, padx=(10, 20), pady=(0, 100))
        self.seasonality_text_frame.pack_propagate(False)
        rec_title = tk.Label(
            self.seasonality_text_frame,
            text="Рекомендации",
            font=("Arial", 14, "bold"),
            fg="#1a5276",
            bg="#f8f9fa"
        )
        rec_title.pack(side=tk.TOP, anchor="w", padx=10, pady=(10, 5))
        self.seasonality_text_widget = tk.Text(
            self.seasonality_text_frame,
            wrap=tk.WORD,
            font=("Arial", 10),
            padx=10,
            pady=10,
            state=tk.DISABLED #Только для чтения
        )
        self.seasonality_text_widget.pack(fill=tk.BOTH, expand=True)
        self.seasonality_bottom_frame = tk.Frame(self.seasonality_frame, bg="#f8f9fa") #Строка статистики для сезонности
        self.seasonality_bottom_frame.pack(fill=tk.X, padx=20)
        self.seasonality_stats_label = tk.Label(
            self.seasonality_bottom_frame,
            text="",
            font=("Arial", 11),
            fg="#2c3e50",
            bg="#f8f9fa",
            justify=tk.LEFT
        )
        self.seasonality_stats_label.pack(side=tk.LEFT)
        self.fig3, self.ax3 = plt.subplots(figsize=(9.5, 4.3)) #График сезонности
        self.canvas3 = FigureCanvasTkAgg(self.fig3, self.seasonality_graph_frame)
        self.canvas3.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5, anchor="nw")
        self.elasticity_frame = tk.Frame(root, bg="#f8f9fa") #Вкладка "Эластичность" (изначально скрыта)
        self.elasticity_top_frame = tk.Frame(self.elasticity_frame, bg="#f8f9fa")
        self.elasticity_top_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
        self.back_from_elasticity_btn = tk.Button(
            self.elasticity_top_frame,
            text="Назад к графикам",
            command=self.show_charts,
            bg="#2c3e50", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5
        )
        self.back_from_elasticity_btn.pack(side=tk.RIGHT, padx=(5, 0))
        elasticity_title = tk.Label(
            self.elasticity_top_frame,
            text="Ценовая эластичность и скидки",
            font=("Arial", 14, "bold"),
            fg="#1a5276",
            bg="#f8f9fa"
        )
        elasticity_title.pack(side=tk.LEFT, padx=10)
        self.elasticity_content_frame = tk.Frame(self.elasticity_frame, bg="#f8f9fa")
        self.elasticity_content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.elasticity_graph_frame = tk.Frame(self.elasticity_content_frame, bg="#f8f9fa")
        self.elasticity_graph_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        self.elasticity_text_frame = tk.Frame(self.elasticity_content_frame, bg="#f8f9fa", width=750, height=100) #Текстовая панель рекомендаций по эластичности
        self.elasticity_text_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 20), pady=(0, 160))
        self.elasticity_text_frame.pack_propagate(False)
        elasticity_rec_title = tk.Label(
            self.elasticity_text_frame,
            text="Рекомендации",
            font=("Arial", 12, "bold"),
            fg="#1a5276",
            bg="#f8f9fa"
        )
        elasticity_rec_title.pack(side=tk.TOP, anchor="w", padx=10, pady=(10, 5))
        self.elasticity_text_widget = tk.Text(
            self.elasticity_text_frame,
            wrap=tk.WORD,
            font=("Arial", 10),
            padx=10,
            pady=10,
            state=tk.DISABLED
        )
        self.elasticity_text_widget.pack(fill=tk.BOTH, expand=True)
        self.elasticity_bottom_frame = tk.Frame(self.elasticity_frame, bg="#f8f9fa") #Строка статистики для эластичности
        self.elasticity_bottom_frame.pack(fill=tk.X, padx=20)
        self.elasticity_stats_label = tk.Label(
            self.elasticity_bottom_frame,
            text="",
            font=("Arial", 11),
            fg="#2c3e50",
            bg="#f8f9fa",
            justify=tk.LEFT
        )
        self.elasticity_stats_label.pack(side=tk.LEFT)
        self.fig_elasticity, self.ax_elasticity = plt.subplots(figsize=(9.5, 4.3)) #График эластичности
        self.canvas_elasticity = FigureCanvasTkAgg(self.fig_elasticity, self.elasticity_graph_frame)
        self.canvas_elasticity.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5, anchor="nw")

    def toggle_theme(self): #Переключение между светлой и тёмной темой интерфейса
        self.dark_theme = not self.dark_theme #Инвертируем текущую тему
        if self.dark_theme: #Определение цветовой палитры
            bg_color = "#2d2d2d" #Цвета для тёмной темы
            fg_color = "white"
            stats_fg = "white"
            canvas_bg = "#1e1e1e"
            legend_bg = "#3a3f4b"
            legend_edge = "#555"
            text_bg = "#2d2d2d"
            text_fg = "white"
        else:
            bg_color = "#f8f9fa" #Цвета для светлой темы (по умолчанию)
            fg_color = "#2c3e50"
            stats_fg = "#2c3e50"
            canvas_bg = "white"
            legend_bg = "white"
            legend_edge = "lightgray"
            text_bg = "white"
            text_fg = "#2c3e50"
        self.root.config(bg=bg_color) #Обновление фона основных фреймов
        self.top_frame.config(bg=bg_color)
        self.control_frame.config(bg=bg_color)
        self.center_container.config(bg=bg_color)
        self.stats_frame.config(bg=bg_color)
        self.charts_frame.config(bg=bg_color)
        self.seasonality_frame.config(bg=bg_color)
        self.seasonality_top_frame.config(bg=bg_color)
        self.seasonality_content_frame.config(bg=bg_color)
        self.seasonality_graph_frame.config(bg=bg_color)
        self.seasonality_text_frame.config(bg=bg_color)
        if hasattr(self, 'seasonality_bottom_frame'):
            self.seasonality_bottom_frame.config(bg=bg_color)
        for widget in self.top_frame.winfo_children() + self.center_container.winfo_children(): #Обновление цветов текстовых меток в верхней панели и центре
            if isinstance(widget, tk.Label):
                widget.config(bg=bg_color, fg=fg_color)
        cat_frame = next( #Обновление фрейма выбора категорий
            (w for w in self.control_frame.winfo_children() if isinstance(w, tk.Frame) and w != self.center_container),
            None
        )
        if cat_frame:
            cat_frame.config(bg=bg_color)
            for widget in cat_frame.winfo_children():
                if isinstance(widget, tk.Label):
                    widget.config(bg=bg_color, fg=fg_color)
        self.theme_btn.config(bg="#7f8c8d" if not self.dark_theme else "#34495e", fg="white") #Обновление цветов кнопок
        self.browse_btn.config(bg="#3498db" if not self.dark_theme else "#2980b9", fg="white")
        self.load_btn.config(bg="#27ae60" if not self.dark_theme else "#218359", fg="white")
        self.export_btn.config(bg="#218359" if not self.dark_theme else "#1a5276", fg="white")
        self.export_plot_btn.config(bg="#21837a" if not self.dark_theme else "#1a5276", fg="white")
        self.seasonality_btn.config(bg="#2f7f99" if not self.dark_theme else "#1a5276", fg="white")
        self.elasticity_btn.config(bg="#2f7f99" if not self.dark_theme else "#1a5276", fg="white")
        for widget in self.seasonality_top_frame.winfo_children(): #Обновление меток во вкладке "Сезонность"
            if isinstance(widget, tk.Label):
                widget.config(bg=bg_color, fg=fg_color)
        for widget in self.seasonality_text_frame.winfo_children():
            if isinstance(widget, tk.Label) and widget.cget("text") == "Рекомендации":
                widget.config(bg=bg_color, fg=fg_color)
        self.stats_label.config(bg=bg_color, fg=stats_fg) #Обновление текста статистики
        if hasattr(self, 'seasonality_stats_label'):
            self.seasonality_stats_label.config(bg=bg_color, fg=stats_fg)
        if hasattr(self, 'seasonality_text_widget'): #Обновление текстовых полей (рекомендаций)
            self.seasonality_text_widget.config(bg=text_bg, fg=text_fg, insertbackground=text_fg)
        for fig, ax in [(self.fig1, self.ax1), (self.fig2, self.ax2)]: #Обновление графиков fig1 и fig2
            fig.patch.set_facecolor(canvas_bg) #Фон холста и области графика
            ax.set_facecolor(canvas_bg)
            ax.tick_params(colors=fg_color) #Цвет меток осей, заголовков и делений
            ax.xaxis.label.set_color(fg_color)
            ax.yaxis.label.set_color(fg_color)
            ax.title.set_color(fg_color)
            for spine in ax.spines.values(): #Цвет рамки графика
                spine.set_color(fg_color)
        for ax in [self.ax1, self.ax2]: #Обновление легенд основных графиков
            legend = ax.get_legend()
            if legend:
                frame = legend.get_frame()
                frame.set_facecolor(legend_bg)
                frame.set_edgecolor(legend_edge)
                frame.set_alpha(0.9)
                for text in legend.get_texts():
                    text.set_color(fg_color)
        if hasattr(self, 'ax1b') and self.ax1b: #Обновление правой оси Y (выручка)
            revenue_color = "white" if self.dark_theme else "black"
            self.ax1b.yaxis.label.set_color(revenue_color)
            self.ax1b.tick_params(axis='y', colors=revenue_color)
            for spine in self.ax1b.spines.values():
                spine.set_color(revenue_color)
        if self.df_full is not None and not self.df_full.empty: #Обновление графика сезонности
            self.fig3.patch.set_facecolor(canvas_bg)
            self.ax3.set_facecolor(canvas_bg)
            self.ax3.tick_params(colors=fg_color)
            self.ax3.xaxis.label.set_color(fg_color)
            self.ax3.yaxis.label.set_color(fg_color)
            self.ax3.title.set_color(fg_color)
            for spine in self.ax3.spines.values():
                spine.set_color(fg_color)
            legend3 = self.ax3.get_legend() #Легенда графика сезонности
            if legend3:
                frame = legend3.get_frame()
                frame.set_facecolor(legend_bg)
                frame.set_edgecolor(legend_edge)
                frame.set_alpha(0.9)
                for text in legend3.get_texts():
                    text.set_color(fg_color)
            self.canvas3.draw()
            if hasattr(self, 'elasticity_frame'): #Обновление вкладки "Эластичность"
                self.elasticity_frame.config(bg=bg_color) #Фоны фреймов
                self.elasticity_top_frame.config(bg=bg_color)
                self.elasticity_content_frame.config(bg=bg_color)
                self.elasticity_graph_frame.config(bg=bg_color)
                self.elasticity_text_frame.config(bg=bg_color)
                if hasattr(self, 'elasticity_bottom_frame'):
                    self.elasticity_bottom_frame.config(bg=bg_color)
                for widget in self.elasticity_top_frame.winfo_children(): #Метки во вкладке эластичности
                    if isinstance(widget, tk.Label):
                        widget.config(bg=bg_color, fg=fg_color)
                if hasattr(self, 'back_from_elasticity_btn'): #Кнопка "Назад"
                    self.back_from_elasticity_btn.config(bg="#2c3e50", fg="white")
                for widget in self.elasticity_text_frame.winfo_children(): #Заголовок "Рекомендации"
                    if isinstance(widget, tk.Label) and widget.cget("text") == "Рекомендации":
                        widget.config(bg=bg_color, fg=fg_color)
                if hasattr(self, 'elasticity_text_widget'): #Текстовое поле рекомендаций
                    self.elasticity_text_widget.config(
                        bg=text_bg,
                        fg=text_fg,
                        insertbackground=text_fg
                    )
                if hasattr(self, 'elasticity_stats_label'): #Статистика эластичности
                    self.elasticity_stats_label.config(bg=bg_color, fg=stats_fg)
            if hasattr(self, 'fig_elasticity') and self.fig_elasticity: #Обновление графика эластичности
                if self.df_full is not None and not self.df_full.empty:
                    self._draw_elasticity_graph()
                else:
                    self.ax_elasticity.clear() #Сообщение при отсутствии данных
                    self.ax_elasticity.text(0.5, 0.5, "Нет данных", ha='center', va='center', fontsize=14,
                                            color='white' if self.dark_theme else 'gray')
                if hasattr(self, '_elasticity_colorbar') and self._elasticity_colorbar: #Цветовая шкала графика эластичности
                    self._elasticity_colorbar.ax.yaxis.set_tick_params(color=fg_color)
                    self._elasticity_colorbar.outline.set_edgecolor(fg_color)
                    plt.setp(self._elasticity_colorbar.ax.get_yticklabels(), color=fg_color)
                    self._elasticity_colorbar.ax.yaxis.label.set_color(fg_color)
                self.canvas_elasticity.draw()
        self.canvas1.draw() #Применение изменений к основным графикам
        self.canvas2.draw()

    def _on_closing(self): #Обработчик закрытия главного окна приложения
        for anim in [self.anim1, self.anim2]: #Остановка анимаций, чтобы избежать утечек памяти
            if anim:
                try:
                    anim.event_source.stop()
                except:
                    pass #Игнорируем ошибки
        try: #Завершение работы главного цикла Tkinter
            self.root.quit() #Останавливает mainloop
            self.root.destroy() #Уничтожает все виджеты
        except:
            pass #Игнорируем ошибки при закрытии

    def browse_file(self): #Открывает диалог выбора Excel-файла и загружает данные для анализа
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")]) #Открытие диалога выбора файла
        if path:
            self.path_entry.delete(0, tk.END) #Очистка и установка пути в поле ввода
            self.path_entry.insert(0, path)
            try:
                df = pd.read_excel(path) #Загрузка данных из Excel
                df.columns = df.columns.str.strip() #Приведение названий столбцов к единому формату
                required = ['Дата', 'Товар', 'Категория', 'Цена за шт', 'Количество продаж', 'Канал продаж', 'Скидки'] #Проверка обязательных столбцов для анализа
                missing = [c for c in required if c not in df.columns]
                if missing:
                    raise ValueError(f"Отсутствуют столбцы: {missing}")
                self.df_full_raw = df.copy() #Сохранение исходных данных
                categories = sorted(df['Категория'].dropna().unique().tolist()) #Формирование списка категорий для выбора пользователем
                self.cat_listbox.delete(1, tk.END) #Обновление списка категорий в интерфейсе
                for cat in categories:
                    self.cat_listbox.insert(tk.END, cat)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось прочитать файл:\n{e}") #Отображение ошибки пользователю

    def apply_category_selection(self): #Обрабатывает выбор категорий пользователем в списке
        indices = self.cat_listbox.curselection() #Получаем индексы выделенных элементов в списке категорий
        if not indices: #Проверка: хотя бы одна категория должна быть выбрана
            messagebox.showwarning("Внимание", "Выберите хотя бы одну категорию.")
            self.cat_listbox.select_set(0) #Автоматически выбираем "Все категории"
            return
        selected = [self.cat_listbox.get(i) for i in indices] #Получаем названия выбранных категорий
        if len(selected) > 5: #Ограничение: не более 5 категорий одновременно
            messagebox.showwarning("Внимание", "Можно выбрать не более 5 категорий.")
            self.cat_listbox.selection_clear(0, tk.END) #Снимаем всё выделение
            for i in indices[:5]: #Восстанавливаем выделение первых 5 позиций
                self.cat_listbox.selection_set(i)
            return
        if "Все категории" in selected: #Особый случай: если выбрана опция "Все категории"
            self.selected_categories = ["Все категории"]
            self.cat_listbox.selection_clear(0, tk.END) #Сбрасываем выделение и выбираем только первую строку
            self.cat_listbox.select_set(0)
        else:
            self.selected_categories = selected #Сохраняем выбранные категории
        if not self.is_loading and self.df_full_raw is not None: #Если анализ не запущен и данные загружены — активируем кнопку анализа
            self.load_btn.config(bg="#27ae60", text="Запустить анализ")

    def start_analysis(self): #Запускает анализ данных в фоновом потоке
        if self.is_loading: #Защита от повторного запуска во время выполнения анализа
            return
        path = self.path_entry.get().strip() #Проверка наличия пути к файлу
        if not path:
            messagebox.showerror("Ошибка", "Укажите путь к файлу.")
            return
        self._set_ui_loading(True) #Блокируем UI и меняем состояние кнопки
        threading.Thread(target=self._background_analysis, args=(path,), daemon=True).start() #Запускаем анализ в отдельном потоке, чтобы не блокировать интерфейс

    def _background_analysis(self, path): #Выполняет полный анализ данных в фоновом потоке
        try:
            self.df_full = None #Сброс предыдущих результатов
            self.monthly_df = None
            self.future_monthly = None
            self.monthly_by_cat = {}
            self.future_by_cat = {}
            df = self.df_full_raw.copy() #Работаем с копией исходных данных
            if "Все категории" in self.selected_categories: #Фильтрация по выбранным категориям
                pass #Используем все данные
            else:
                df = df[df['Категория'].isin(self.selected_categories)].copy()
                if df.empty:
                    raise ValueError(f"Нет данных по выбранным категориям: {self.selected_categories}")
            df['Дата'] = pd.to_datetime(df['Дата'], dayfirst=True, errors='coerce') #Преобразование даты
            df = df.dropna(subset=['Дата'])
            df['Цена за шт'] = pd.to_numeric(df['Цена за шт'], errors='coerce') #Преобразование числовых столбцов
            df['Количество продаж'] = pd.to_numeric(df['Количество продаж'], errors='coerce')
            df['Скидки'] = df['Скидки'].fillna('нет').astype(str).str.strip().str.lower() #Обработка скидок: приведение к числовому виду
            df['Скидки'] = df['Скидки'].replace({'нет': 0, '': 0, '-': 0})
            df['Скидки'] = pd.to_numeric(df['Скидки'], errors='coerce').fillna(0)
            df = df.dropna(subset=['Цена за шт', 'Количество продаж']) #Удаление строк с отсутствующими ценой или количеством
            if len(df) < 5: #Проверка минимального объёма данных для анализа
                raise ValueError(f"Слишком мало данных (требуется ≥5 записей, имеется: {len(df)})")
            self.df_full = df.copy() #Сохраняем обработанный датафрейм
            le_cat = LabelEncoder() #Кодирование категориальных признаков
            le_channel = LabelEncoder()
            df['код_категории'] = le_cat.fit_transform(df['Категория'])
            df['код_канала'] = le_channel.fit_transform(df['Канал продаж'])
            df['месяц'] = df['Дата'].dt.month #Генерация временных признаков
            df['квартал'] = df['Дата'].dt.quarter
            df['день_недели'] = df['Дата'].dt.dayofweek
            df['выходной'] = (df['день_недели'] >= 5).astype(int)
            numeric_cols = ['Цена за шт', 'Скидки', 'месяц', 'квартал', 'день_недели', 'выходной'] #Подготовка матрицы признаков X и целевой переменной y
            categorical_cols = ['код_категории', 'код_канала']
            scaler = StandardScaler() #Нормализация числовых признаков
            X_numeric = scaler.fit_transform(df[numeric_cols])
            X_categorical = df[categorical_cols].values
            X = np.hstack([X_numeric, X_categorical])
            y = df['Количество продаж'].values
            model = None
            try: #Обучение модели прогнозирования спроса
                model = NNBoost(n_estimators=50, learning_rate=0.03)
                model.fit(X, y)
                pred = model.predict(X)
                df['прогноз_спроса'] = pred
                df['оценка_выручки'] = df['Цена за шт'] * df['прогноз_спроса']
            except Exception as e_inner: #Если обучение не удалось, используем фактические данные как прогноз
                print(f"Обучение не удалось: {e_inner}. Fallback.")
                df['прогноз_спроса'] = df['Количество продаж']
                df['оценка_выручки'] = df['Количество продаж'] * df['Цена за шт']
            if 'оценка_выручка' not in df.columns: #Гарантируем наличие колонок прогноза
                df['оценка_выручка'] = df['Цена за шт'] * df['Количество продаж']
            if 'прогноз_спроса' not in df.columns:
                df['прогноз_спроса'] = df['Количество продаж']
            self.df_full = df.copy()
            df['месяц_period'] = df['Дата'].dt.to_period('M') #Агрегация по месяцам
            monthly = df.groupby('месяц_period', as_index=False).agg({
                'Количество продаж': 'sum',
                'прогноз_спроса': 'sum',
                'оценка_выручки': 'sum',
                'Цена за шт': 'mean'
            }).reset_index(drop=True)
            for col in ['прогноз_спроса', 'оценка_выручка']: #Заполняем недостающие колонки
                if col not in monthly.columns:
                    monthly[col] = monthly['Количество продаж'] * (monthly['Цена за шт'] if col == 'оценка_выручка' else 1)
            monthly['месяц_str'] = monthly['месяц_period'].astype(str)
            monthly = monthly.sort_values('месяц_period').reset_index(drop=True)
            self.monthly_df = monthly
            self.price_elasticity, self.discount_effect, elasticity_insight = self._calculate_price_elasticity_and_discount_effect(
                df) #Расчёт ценовой эластичности и эффекта скидок
            if "Все категории" not in self.selected_categories and len(self.selected_categories) > 1: #Прогноз на 2026 год
                for cat in self.selected_categories: #Прогноз по каждой выбранной категории отдельно
                    cat_df = df[df['Категория'] == cat].copy()
                    if 'прогноз_спроса' not in cat_df.columns: #Гарантируем наличие колонок прогноза для категории
                        cat_df['прогноз_спроса'] = cat_df['Количество продаж']
                    if 'оценка_выручка' not in cat_df.columns:
                        cat_df['оценка_выручка'] = cat_df['Количество продаж'] * cat_df['Цена за шт']
                    if cat_df.empty: #Если данных по категории нет, заполняем нулями
                        self.monthly_by_cat[cat] = pd.DataFrame({
                            'месяц_period': monthly['месяц_period'],
                            'Количество продаж': np.zeros(len(monthly)),
                            'оценка_выручка': np.zeros(len(monthly))
                        }).set_index('месяц_period')
                        self.future_by_cat[cat] = np.zeros(12)
                        continue
                    cat_monthly = cat_df.groupby('месяц_period').agg({ #Месячная агрегация по категории
                        'Количество продаж': 'sum',
                        'оценка_выручка': 'sum'
                    })
                    cat_monthly = cat_monthly.reindex(monthly['месяц_period']).fillna(0)
                    self.monthly_by_cat[cat] = cat_monthly
                    try: #Прогноз на 12 месяцев вперёд
                        future_rows = []
                        last_date = cat_df['Дата'].max()
                        for i in range(1, 13):
                            next_date = last_date + pd.DateOffset(months=i)
                            hist = cat_df[cat_df['Дата'].dt.month == next_date.month] #Берём исторические данные за тот же месяц
                            if len(hist) == 0:
                                hist = cat_df.sample(min(50, len(cat_df)), replace=True) #Если нет данных, сэмплируем из всего набора
                            for _, row in hist.iterrows():
                                new_price = row['Цена за шт'] * (1.005 ** i) #Моделируем небольшой рост цены
                                future_rows.append({
                                    'Цена за шт': new_price,
                                    'Скидки': row['Скидки'],
                                    'месяц': next_date.month,
                                    'квартал': next_date.quarter,
                                    'день_недели': 0,
                                    'выходной': 0,
                                    'код_категории': row['код_категории'],
                                    'код_канала': row['код_канала']
                                })
                        future_df = pd.DataFrame(future_rows)
                        X_fut = np.hstack([  #Подготовка признаков для прогноза
                            scaler.transform(future_df[numeric_cols]),
                            future_df[categorical_cols].values
                        ])
                        future_df['прогноз_спроса'] = model.predict(X_fut) #Прогноз спроса и выручки
                        future_df['прогноз_выручка'] = future_df['Цена за шт'] * future_df['прогноз_спроса']
                        rev_series = future_df.groupby(future_df.index // max(1, len(future_df)//12))['прогноз_выручка'].sum() #Агрегация по месяцам
                        vals = rev_series.values[:12]
                        if len(vals) < 12:
                            vals = np.pad(vals, (0, 12 - len(vals)), constant_values=(vals[-1] if len(vals) > 0 else 50000))
                        self.future_by_cat[cat] = vals
                    except Exception as e_inner: #Fallback для прогноза по категории
                        print(f"Прогноз для {cat} не удался: {e_inner}")
                        avg_monthly = cat_df.groupby(cat_df['Дата'].dt.to_period('M'))['Количество продаж'].sum().mean()
                        avg_price = cat_df['Цена за шт'].mean()
                        self.future_by_cat[cat] = np.full(12, avg_monthly * avg_price)
            else: #Прогноз для "Все категории" или одной категории
                future_rows = []
                last_date = df['Дата'].max()
                for i in range(1, 13):
                    next_date = last_date + pd.DateOffset(months=i)
                    hist = df[df['Дата'].dt.month == next_date.month]
                    if len(hist) == 0:
                        hist = df.sample(min(50, len(df)), replace=True)
                    for _, row in hist.iterrows():
                        new_price = row['Цена за шт'] * (1.005 ** i)
                        future_rows.append({
                            'Цена за шт': new_price,
                            'Скидки': row['Скидки'],
                            'месяц': next_date.month,
                            'квартал': next_date.quarter,
                            'день_недели': 0,
                            'выходной': 0,
                            'код_категории': row['код_категории'],
                            'код_канала': row['код_канала']
                        })
                future_df = pd.DataFrame(future_rows)
                X_fut = np.hstack([
                    scaler.transform(future_df[numeric_cols]),
                    future_df[categorical_cols].values
                ])
                try:
                    future_df['прогноз_спроса'] = model.predict(X_fut)
                except: #Fallback: если модель недоступна
                    future_df['прогноз_спроса'] = future_df['Цена за шт']
                future_df['прогноз_выручка'] = future_df['Цена за шт'] * future_df['прогноз_спроса']
                grouped = future_df.groupby(future_df.index // max(1, len(future_df)//12)) #Агрегация по месяцам
                future_monthly = pd.DataFrame({
                    'month_id': range(12),
                    'прогноз_выручка': grouped['прогноз_выручка'].sum().values[:12],
                    'прогноз_спроса': grouped['прогноз_спроса'].sum().values[:12]
                })
                while len(future_monthly) < 12: #Дополнение до 12 месяцев, если данных не хватает
                    future_monthly = pd.concat([future_monthly, pd.DataFrame([{
                        'month_id': len(future_monthly),
                        'прогноз_выручка': future_monthly['прогноз_выручка'].mean(),
                        'прогноз_спроса': future_monthly['прогноз_спроса'].mean()
                    }])], ignore_index=True)
                future_monthly = future_monthly.head(12)
                future_monthly['месяц_str'] = [(last_date + pd.DateOffset(months=i+1)).strftime('%Y-%m') for i in range(12)]
                self.future_monthly = future_monthly
            self.root.after(0, self._update_ui_with_results, self.selected_categories) #Обновление UI в основном потоке
        except Exception as e: #Обработка ошибок и вывод пользователю
            import traceback
            err_msg = str(e)
            print("Ошибка в фоновом потоке:", err_msg)
            print(traceback.format_exc())
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("Ошибка", msg))
        finally: #Восстановление состояния UI после завершения анализа
            self.root.after(0, self._set_ui_loading, False)

    def _update_ui_with_results(self, categories): #Обновляет пользовательский интерфейс после завершения анализа данных
        try: #Расчёт сезонности (даже если не отображается сразу)
            seasonality_text = ""
            if self.df_full is not None and not self.df_full.empty:
                seasonal_coeffs, seasonality_text = self._calculate_seasonality(self.df_full)
                self.seasonal_coeffs = seasonal_coeffs
            if self.current_view == "seasonality": #Обновление текущего вида (графики/сезонность/эластичность)
                self._draw_seasonality_graph()
            elif self.current_view == "elasticity":
                self._draw_elasticity_graph()
            else:
                self._draw_animation(categories)
            if self.df_full is not None and not self.df_full.empty: #Формирование строки статистики под графиками
                avg_qty = self.df_full['Количество продаж'].mean()
                avg_price = self.df_full['Цена за шт'].mean()
                total_rev_exact = (self.df_full['Количество продаж'] * self.df_full['Цена за шт']).sum()
                if self.future_monthly is not None: #Прогноз выручки на 2026 год
                    forecast_2026 = self.future_monthly['прогноз_выручка'].sum()
                elif self.future_by_cat:
                    forecast_2026 = sum(arr.sum() for arr in self.future_by_cat.values())
                else:
                    forecast_2026 = 0
                metrics, years_info = self._calculate_yearly_comparison(self.df_full) #Расчёт изменений относительно предыдущего года

                def fmt_int(x): #Форматирование целых чисел с пробелами вместо запятых
                    return f"{int(round(x)):,}".replace(",", " ")

                def fmt_pct(change): #Форматирование процентного изменения со стрелками
                    if abs(change) < 0.1:
                        return "≈0%"
                    arrow = "▲" if change > 0 else "▼"
                    return f"{arrow}{abs(change):.0f}%"
                qty_str = f"{fmt_int(avg_qty)} шт. ({fmt_pct(metrics.get('qty_change', 0))})" #Формирование строк для отображения
                price_str = f"{fmt_int(avg_price)} ₽ ({fmt_pct(metrics.get('price_change', 0))})"
                rev_str = f"{fmt_int(total_rev_exact)} ₽ ({fmt_pct(metrics.get('rev_change', 0))})"
                stats_parts = [
                    f"Средний спрос: {qty_str}",
                    f"Средняя цена: {price_str}",
                    f"Фактич. выручка: {rev_str}",
                    f"Прогноз 2026: {fmt_int(forecast_2026)} ₽"
                ]
                stats_text = "  •  ".join(stats_parts)
                self.stats_label.config(text=stats_text)
            self.export_btn.config(state="normal") #Активация кнопок экспорта после успешного анализа
            self.export_plot_btn.config(state="normal")
            self.seasonality_btn.config(state="normal")
        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("Ошибка визуализации", msg))

    def _set_ui_loading(self, state): #Блокирует/разблокирует элементы управления во время выполнения анализа
        self.is_loading = state
        if state: #Блокировка всех элементов управления во время анализа
            self.load_btn.config(state="disabled", text="Анализ...")
            self.browse_btn.config(state="disabled")
            self.apply_cat_btn.config(state="disabled")
            self.cat_listbox.config(state="disabled")
            self.export_btn.config(state="disabled")
            self.elasticity_btn.config(state="disabled")
        else: #Восстановление состояния элементов после завершения анализа
            self.load_btn.config(state="normal", text="Запустить анализ")
            self.browse_btn.config(state="normal")
            self.apply_cat_btn.config(state="normal")
            self.cat_listbox.config(state="normal")
            self.elasticity_btn.config(state="normal" if self.df_full is not None else "disabled")

    def _calculate_yearly_comparison(self, df): #Рассчитывает сравнение показателей текущего года с предыдущим
        if df.empty or 'Дата' not in df.columns:
            return {}, {}
        df['год'] = df['Дата'].dt.year  #Извлечение года из даты
        years = sorted(df['год'].unique())
        if len(years) < 2:
            return {}, {} #Недостаточно данных для сравнения
        current_year = years[-1]
        prev_year = years[-2]
        yearly = df.groupby('год').agg( #Агрегация данных по годам
            total_qty=('Количество продаж', 'sum'),
            avg_price=('Цена за шт', 'mean'),
            total_rev=('оценка_выручка', 'sum')
        ).round(2)
        curr = yearly.loc[current_year]
        prev = yearly.loc[prev_year]

        def pct_change(curr, prev): #Расчёт процентного изменения
            if prev == 0:
                return 0
            return (curr - prev) / prev * 100
        metrics = {
            'avg_qty': df[df['год'] == current_year]['Количество продаж'].mean(),
            'avg_qty_prev': df[df['год'] == prev_year]['Количество продаж'].mean(),
            'avg_price': curr['avg_price'],
            'avg_price_prev': prev['avg_price'],
            'total_rev': curr['total_rev'],
            'total_rev_prev': prev['total_rev']
        }
        metrics['qty_change'] = pct_change(metrics['avg_qty'], metrics['avg_qty_prev']) #Процентные изменения
        metrics['price_change'] = pct_change(metrics['avg_price'], metrics['avg_price_prev'])
        metrics['rev_change'] = pct_change(metrics['total_rev'], metrics['total_rev_prev'])
        return metrics, {
            'current_year': current_year,
            'prev_year': prev_year
        }

    def _calculate_seasonality(self, df): #Рассчитывает коэффициенты сезонности по месяцам для каждой категории
        if df.empty or 'Дата' not in df.columns:
            return {}, ""
        df = df.copy()
        df['месяц'] = df['Дата'].dt.month
        df['Категория'] = df['Категория'].fillna('Прочее')
        grouped = df.groupby(['Категория', 'месяц'])['Количество продаж'].sum().reset_index() #Группировка по категории и месяцу: суммарный спрос
        seasonal_coeffs = {}
        insights = []
        for cat in grouped['Категория'].unique():
            cat_data = grouped[grouped['Категория'] == cat]
            avg_qty = cat_data['Количество продаж'].mean()
            if avg_qty == 0:
                continue
            coeffs = { #Расчёт коэффициента сезонности для каждого месяца
                int(row['месяц']): row['Количество продаж'] / avg_qty
                for _, row in cat_data.iterrows()
            }
            if max(coeffs.values()) > 1.7: # Сохранение только если есть выраженный пик (>1.7x)
                seasonal_coeffs[cat] = coeffs
                peak_month, peak_val = max(coeffs.items(), key=lambda x: x[1]) #Определяем месяц с максимальным спросом
                month_name = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
                              'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек'][peak_month - 1]
                insights.append(f"{month_name} ×{peak_val:.1f} ({cat})")
        insight_str = " • ".join(insights[:3]) #Формируем краткую строку с топ-3 пиками
        if len(insights) > 3:
            insight_str += " • ..."
        return seasonal_coeffs, insight_str

    def _generate_seasonality_recommendations(self): #Генерирует бизнес-рекомендации на основе выявленной сезонности
        if not hasattr(self, 'seasonal_coeffs') or not self.seasonal_coeffs:
            return "Чёткой сезонности не обнаружено.\nРекомендуется собрать больше данных (≥2 года)."
        month_names = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                       "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
        recs = []
        for cat, coeffs in self.seasonal_coeffs.items():
            peaks = [m for m, c in coeffs.items() if c > 1.8] #Месяцы с высоким спросом
            lows = [m for m, c in coeffs.items() if c < 0.6] #Месяцы с низким спросом
            if peaks:
                peak_months = ", ".join([month_names[m - 1] for m in sorted(peaks)])
                recs.append(f"• {cat}: повышенный спрос в {peak_months} → увеличить закупки за 1–2 мес.")
            if lows:
                low_months = ", ".join([month_names[m - 1] for m in sorted(lows)])
                recs.append(f"• {cat}: низкий спрос в {low_months} → избегать закупок, запускать акции.")
        if not recs:
            return "Чёткой сезонности не обнаружено.\nРекомендуется собрать больше данных (≥2 года)."
        return "\n".join(recs)

    def _calculate_price_elasticity_and_discount_effect(self, df): #Рассчитывает ценовую эластичность спроса по категориям и эффект от скидок
        if df.empty or 'Скидки' not in df.columns:
            return {}, {}, ""
        df = df.copy()
        df['Скидки'] = pd.to_numeric(df['Скидки'], errors='coerce').fillna(0)
        df['Цена со скидкой'] = df['Цена за шт'] * (1 - df['Скидки'] / 100)
        elasticity = {} #Категория: эластичность
        discount_effect = {} #Категория: спрос_% и выручки_%
        insights = [] #Текстовые выводы по эффективности скидок
        for cat in df['Категория'].dropna().unique():
            cat_df = df[df['Категория'] == cat]
            if len(cat_df) < 10: #Недостаточно данных для анализа
                continue
            cat_df = cat_df.sort_values('Дата')
            cat_df['price_pct'] = cat_df['Цена за шт'].pct_change() #Относительные изменения цены и спроса
            cat_df['qty_pct'] = cat_df['Количество продаж'].pct_change()
            valid = cat_df[(cat_df['price_pct'].abs() > 0.01) & (cat_df['qty_pct'].abs() > 0.01)] #Отбираем значимые изменения
            if len(valid) > 5:
                el = (valid['qty_pct'] / valid['price_pct']).median() #Медианная эластичность
                elasticity[cat] = -el #Инвертируем знак для удобства интерпретации
            with_disc = cat_df[cat_df['Скидки'] > 0] #Анализ эффекта скидок: сравнение периодов со скидками и без
            without_disc = cat_df[cat_df['Скидки'] == 0]
            if len(with_disc) > 5 and len(without_disc) > 5:
                qty_w = with_disc['Количество продаж'].mean()
                qty_wo = without_disc['Количество продаж'].mean()
                rev_w = (with_disc['Цена со скидкой'] * with_disc['Количество продаж']).mean()
                rev_wo = (without_disc['Цена за шт'] * without_disc['Количество продаж']).mean()
                delta_qty = (qty_w - qty_wo) / qty_wo * 100 if qty_wo > 0 else 0
                delta_rev = (rev_w - rev_wo) / rev_wo * 100 if rev_wo > 0 else 0
                discount_effect[cat] = (delta_qty, delta_rev)
                if delta_rev > 5: #Формирование инсайтов по эффективности скидок
                    insights.append(f"Скидки эффективны: {cat} (+{delta_rev:.0f}% выручки)")
                elif delta_rev < -5:
                    insights.append(f"⚠Скидки вредят: {cat} ({delta_rev:.0f}% выручки)")
        insight_str = " • ".join(insights[:2]) #Краткая строка с топ-2 инсайтами
        if len(insights) > 2:
            insight_str += " • ..."
        return elasticity, discount_effect, insight_str

    def _generate_elasticity_recommendations(self): #Генерирует рекомендации по ценообразованию и использованию скидок
        if not hasattr(self, 'price_elasticity') or not self.price_elasticity:
            return "Недостаточно данных для анализа эластичности."
        recs = []
        for cat, el in self.price_elasticity.items():
            if el < -1.5:
                recs.append(f"• {cat}: высокая эластичность ({el:.2f}) → снижайте цены для роста выручки.")
            elif -1.5 <= el < -1.0:
                recs.append(f"• {cat}: умеренная эластичность ({el:.2f}) → осторожное ценообразование.")
            elif -1.0 <= el <= -0.5:
                recs.append(f"• {cat}: низкая эластичность ({el:.2f}) → можно повышать цены.")
            elif el > -0.5:
                recs.append(f"• {cat}: неэластичный спрос ({el:.2f}) → цены не влияют на объёмы.")
        if not recs:
            return "Эластичность не определена."
        if hasattr(self, 'discount_effect'): #Дополнительные рекомендации по скидкам
            for cat, (delta_qty, delta_rev) in self.discount_effect.items():
                if delta_rev > 5:
                    recs.append(f"• Скидки для '{cat}' эффективны (+{delta_rev:.1f}% выручки).")
                elif delta_rev < -3:
                    recs.append(f"• Скидки для '{cat}' вредят выручке ({delta_rev:.1f}%). Пересмотрите условия.")
        return "\n".join(recs)

    def _draw_animation(self, categories): #Отрисовка графиков с анимацией прозрачности
        for anim in [self.anim1, self.anim2]: #Остановка предыдущих анимаций
            if anim:
                try:
                    anim.event_source.stop()
                except:
                    pass
        self.anim1 = self.anim2 = None
        self.fig1.clear() #Очистка и подготовка холстов
        self.fig2.clear()
        self.ax1 = self.fig1.add_subplot(111)
        self.ax2 = self.fig2.add_subplot(111)
        self.ax1b = self.ax1.twinx() #Вторая ось Y для выручки на первом графике
        revenue_color = "white" if self.dark_theme else "black"
        self.ax1b.yaxis.label.set_color(revenue_color)
        self.ax1b.tick_params(axis='y', colors=revenue_color)
        for spine in self.ax1b.spines.values():
            spine.set_color(revenue_color)
        self.anim_bars = []  #Списки для управления анимацией элементов
        self.anim_lines1 = []
        self.canvas1.figure = self.fig1 #Привязка фигур к холстам
        self.canvas2.figure = self.fig2
        if self.dark_theme: #Настройка цветовой палитры в зависимости от темы
            canvas_bg = "#1e1e1e"
            text_color = "white"
            legend_bg = "#3a3f4b"
            legend_edge = "#555"
            legend_text = "white"
        else:
            canvas_bg = "white"
            text_color = "#2c3e50"
            legend_bg = "white"
            legend_edge = "lightgray"
            legend_text = "#2c3e50"
        self.fig1.patch.set_facecolor(canvas_bg) #Применение цветов к холстам и осям
        self.ax1.set_facecolor(canvas_bg)
        self.fig2.patch.set_facecolor(canvas_bg)
        self.ax2.set_facecolor(canvas_bg)
        for ax in [self.ax1, self.ax1b, self.ax2]:
            ax.tick_params(colors=text_color)
            ax.xaxis.label.set_color(text_color)
            ax.yaxis.label.set_color(text_color)
            ax.title.set_color(text_color)
            for spine in ax.spines.values():
                spine.set_color(text_color)
        df = self.monthly_df #Проверка наличия данных
        if df is None or df.empty:
            self.ax1.text(0.5, 0.5, "Нет данных", ha='center', va='center', fontsize=14, color='gray')
            self.ax2.text(0.5, 0.5, "Нет данных", ha='center', va='center', fontsize=14, color='gray')
            self.canvas1.draw()
            self.canvas2.draw()
            return
        n = len(df) #Подготовка данных для отрисовки
        x = np.arange(n)
        x_labels = df['месяц_str'].tolist() if 'месяц_str' in df.columns else [f"M{i + 1}" for i in range(n)]
        title = "все категории" if "Все категории" in categories else ", ".join(categories)
        if "Все категории" in categories or len(categories) == 1: #Отрисовка первого графика, режим "Все категории" или одна категория: простые столбцы и линия
            bars = self.ax1.bar(x, df['Количество продаж'], width=0.6, color='#76c68f', alpha=0.0, label='Факт: спрос')
            line, = self.ax1b.plot(x, df['оценка_выручка'], 'o-', color='#22a7f0', linewidth=2.2, markersize=6,
                              alpha=0.0, label='Оценка: выручка')
            self.anim_bars = [bars]
            self.anim_lines1 = [line]
        else: #Режим нескольких категорий: стековые столбцы
            bottom = np.zeros(n)
            demand_colors = ['#3eb489', '#6db0cd', '#9ec0de', '#a3de9e', '#96afdb']
            revenue_colors = ['#3e7cb4', '#2c6c7e', '#3e55b4', '#1a414c', '#11442e']
            for i, cat in enumerate(categories):
                color_demand = demand_colors[i % len(demand_colors)]
                color_revenue = revenue_colors[i % len(revenue_colors)]
                if cat in self.monthly_by_cat: #Получение данных по категории
                    series = self.monthly_by_cat[cat]
                    qty = series['Количество продаж'].values
                    rev = series['оценка_выручка'].values
                else:
                    qty = np.zeros(n)
                    rev = np.zeros(n)
                bars = self.ax1.bar(x, qty, bottom=bottom, width=0.6, #Отрисовка стекового столбца
                                    color=color_demand, alpha=0.0,
                                    label=f'{cat}: факт')
                for rect in bars: #Сохранение категории в каждом прямоугольнике для подсказок
                    rect._demand_category = cat
                line, = self.ax1b.plot(x, rev, '^-', color=color_revenue, #Отрисовка линии выручки
                                  linewidth=1.8, markersize=5,
                                  alpha=0.0, label=f'{cat}: выручка')
                self.anim_bars.append(bars)
                self.anim_lines1.append(line)
                bottom += qty
        self.ax1.set_ylabel('Спрос, шт', fontsize=10) #Настройка осей и оформления первого графика
        self.ax1b.set_ylabel('Выручка, ₽', fontsize=10)
        self.ax1.set_xlabel('Месяц', fontsize=10)
        self.ax1.set_title(f'Факт и прогноз спроса\n({title})', fontweight='bold', fontsize=11)
        self.ax1.set_xticks(x)
        self.ax1.set_xticklabels(x_labels, rotation=45, ha='right', fontsize=8)
        self.ax1.grid(True, linestyle='--', alpha=0.5, linewidth=0.7)
        self.ax1.spines['top'].set_visible(False)
        self.ax1b.spines['top'].set_visible(False)
        self.ax1.set_ylim(bottom=0)
        self.ax1b.set_ylim(bottom=0)
        self.ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x):,}'.replace(',', ' '))) #Форматирование чисел с пробелами вместо запятых
        self.ax1b.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x):,}'.replace(',', ' ')))
        handles1, labels1 = self.ax1.get_legend_handles_labels() #Создание легенды для первого графика
        handles2, labels2 = self.ax1b.get_legend_handles_labels()
        legend_elements = []
        if "Все категории" in categories or len(categories) == 1: #Легенда для режима "Все категории"
            bars = self.ax1.containers[0]
            rect = bars[0]
            fc = rect.get_facecolor()
            legend_elements.append(Patch(facecolor=fc, alpha=0.8, label='Факт: спрос'))
            line = self.ax1b.get_lines()[0]
            color = line.get_color()
            legend_elements.append(Line2D([0], [0], color=color, marker='o', linestyle='-',
                                          linewidth=2.2, markersize=6, label='Оценка: выручка'))
        else: #Легенда для режима нескольких категорий
            for i, cat in enumerate(categories):
                if i < len(self.ax1.containers):
                    bars = self.ax1.containers[i]
                    rect = bars[0]
                    fc = rect.get_facecolor()
                    legend_elements.append(Patch(facecolor=fc, alpha=0.7, label=f'{cat}: факт'))
                if i < len(self.ax1b.get_lines()):
                    line = self.ax1b.get_lines()[i]
                    color = line.get_color()
                    legend_elements.append(Line2D([0], [0], color=color, marker='^', linestyle='-',
                                                  linewidth=1.8, markersize=5, label=f'{cat}: выручка'))
        legend = self.ax1.legend(handles=legend_elements, loc='upper left', #Настройка легенды
                                 bbox_to_anchor=(0.0, -0.27),
                                 fontsize=8,
                                 ncol=1 if len(legend_elements) < 8 else 2,
                                 frameon=True)
        legend.get_frame().set_facecolor(legend_bg)
        legend.get_frame().set_edgecolor(legend_edge)
        legend.get_frame().set_alpha(0.9)
        for text in legend.get_texts():
            text.set_color(legend_text)
        for artist in legend.get_children():
            artist.set_animated(False)
        x_vals = np.arange(12) #Отрисовка второго графика
        month_labels = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
        self.anim_lines2 = []
        revenue_colors = ['#3e7cb4', '#2c6c7e', '#3e55b4', '#1a414c', '#11442e']
        if "Все категории" in categories or len(categories) == 1: #Прогноз для всех категорий
            y = self.future_monthly['прогноз_выручка'].values[:12].astype(float)
            if len(y) < 12:
                y = np.pad(y, (0, 12 - len(y)), constant_values=(y[-1] if len(y) > 0 else 50000))
            line, = self.ax2.plot(x_vals, y, 's-', color='#a6d75b', linewidth=2.5, markersize=7,
                                  alpha=0.0, label='2026 (суммарно)')
            self.anim_lines2 = [line]
        else: #Прогноз по категориям
            for i, cat in enumerate(categories):
                color = revenue_colors[i % len(revenue_colors)]
                if cat in self.future_by_cat and len(self.future_by_cat[cat]) >= 12:
                    y = self.future_by_cat[cat][:12]
                else:
                    y = np.zeros(12)
                line, = self.ax2.plot(x_vals, y, 'o--', color=color, linewidth=2, markersize=5,
                                      alpha=0.0, label=cat)
                self.anim_lines2.append(line)
        self.ax2.set_title(f'Прогноз выручки на 2026\n({title})', fontweight='bold', fontsize=11) #Настройка второго графика
        self.ax2.set_xlabel('Месяц', fontsize=10)
        self.ax2.set_ylabel('Выручка (₽)', fontsize=10)
        self.ax2.set_xticks(x_vals)
        self.ax2.set_xticklabels(month_labels, fontsize=9)
        self.ax2.grid(True, linestyle='--', alpha=0.5, linewidth=0.7)
        self.ax2.spines['top'].set_visible(False)
        self.ax2.spines['right'].set_visible(False)
        self.ax2.set_ylim(bottom=0)
        self.ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{int(x):,}'.replace(',', ' ')))
        revenue_colors = ['#3e7cb4', '#2c6c7e', '#3e55b4', '#1a414c', '#11442e']
        legend2_elements = [] #Легенда для второго графика
        if "Все категории" in categories or len(categories) == 1:
            legend2_elements.append(Line2D([0], [0], color='#a6d75b', marker='s', linestyle='-',
                                           linewidth=2.5, markersize=7, label='2026 (суммарно)'))
        else:
            for i, cat in enumerate(categories):
                color = revenue_colors[i % len(revenue_colors)]
                legend2_elements.append(Line2D([0], [0], color=color, marker='o', linestyle='--',
                                               linewidth=2, markersize=5, label=cat))
        legend2 = self.ax2.legend(handles=legend2_elements, loc='upper left',
                                  bbox_to_anchor=(0.0, -0.27),
                                  fontsize=8,
                                  ncol=1 if len(legend2_elements) < 5 else 2,
                                  frameon=True)
        legend2.get_frame().set_facecolor(legend_bg)
        legend2.get_frame().set_edgecolor(legend_edge)
        legend2.get_frame().set_alpha(0.9)
        for text in legend2.get_texts():
            text.set_color(legend_text)
        for artist in legend2.get_children():
            artist.set_animated(False)
        legend.set_zorder(1000) #Установка приоритета отображения легенд
        legend2.set_zorder(1000)
        self.fig1.subplots_adjust(bottom=0.37) #Настройка отступов для размещения легенд
        self.fig2.subplots_adjust(bottom=0.37)
        self.canvas1.draw() #Первичная отрисовка графиков
        self.canvas2.draw()
        steps = 10
        delay = 60

        def animate_step(step): #Постепенное увеличение прозрачности элементов графика
            alpha = step / steps
            for bars in self.anim_bars:
                for rect in bars:
                    rect.set_alpha(alpha)
            for line in self.anim_lines1:
                line.set_alpha(alpha)
            for line in self.anim_lines2:
                line.set_alpha(alpha)
            self.canvas1.draw_idle()
            self.canvas2.draw_idle()
            if step < steps:
                self.root.after(delay, animate_step, step + 1)
            else:
                pass
        self.root.after(100, animate_step, 1)
        if hasattr(self, '_cursor_fig1'): #Настройка интерактивных подсказок для первого графика
            self._cursor_fig1.remove()
        artists = []
        for container in self.ax1.containers:
            artists.extend(container)
        artists.extend(self.ax1b.get_lines())
        self._cursor_fig1 = mplcursors.cursor(artists, hover=True) #Настройка интерактивных подсказок при наведении
        x_labels = df['месяц_str'].tolist() if 'месяц_str' in df.columns else [f"M{i + 1}" for i in range(len(df))]
        @self._cursor_fig1.connect("add")

        def on_add(sel):
            sel.annotation.set( #Стиль подсказки: белый фон, чёрная рамка, непрозрачный
                bbox=dict(boxstyle="round,pad=0.5", fc="white", ec="black", alpha=1.0),
                color="black",
                fontsize=9
            )
            if "Все категории" in categories or len(categories) == 1: #Определение категории для отображения
                category = "Суммарно"
            else:
                if isinstance(sel.artist, mpl.patches.Rectangle):
                    category = getattr(sel.artist, '_demand_category', 'Неизвестно') #Извлечение категории из сохранённого атрибута
                else:
                    label = sel.artist.get_label() #Для линий — из метки (label)
                    category = label.replace(": выручка", "").replace(": факт", "").strip()
            if isinstance(sel.artist, mpl.patches.Rectangle):  #Формирование текста подсказки
                x_center = sel.artist.get_x() + sel.artist.get_width() / 2
                idx = int(round(x_center))
                if 0 <= idx < len(x_labels):
                    month = x_labels[idx]
                    value = int(sel.artist.get_height())
                    sel.annotation.set_text(
                        f"Категория: {category}\nМесяц: {month}\nСпрос: {value:,} шт".replace(',', ' ')
                    )
                else:
                    sel.annotation.set_visible(False)
                    self.canvas1.draw_idle()
            elif isinstance(sel.artist, mpl.lines.Line2D):
                x_data = sel.target[0]
                idx = int(round(x_data))
                if 0 <= idx < len(x_labels):
                    month = x_labels[idx]
                    value = int(round(sel.target[1]))
                    sel.annotation.set_text(
                        f"Категория: {category}\nМесяц: {month}\nВыручка: {value:,} ₽".replace(',', ' ')
                    )
                else:
                    sel.annotation.set_visible(False)
                    self.canvas1.draw_idle()
            if hasattr(self, '_tooltip_hide_timer'): #Таймер автоматического скрытия подсказки через 1 секунду
                self.root.after_cancel(self._tooltip_hide_timer)
            self._tooltip_hide_timer = self.root.after(1000, lambda s=sel: self._hide_tooltip_safe(s)) # Запускаем новый
        if hasattr(self, '_cursor_fig2'): #Настройка интерактивных подсказок для второго графика
            self._cursor_fig2.remove()
        artists_fig2 = []
        artists_fig2.extend(self.ax2.get_lines())
        self._cursor_fig2 = mplcursors.cursor(artists_fig2, hover=True)
        month_labels = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                        "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
        @self._cursor_fig2.connect("add")

        def on_add_fig2(sel):
            sel.annotation.set(
                bbox=dict(boxstyle="round,pad=0.5", fc="white", ec="black", alpha=1.0),
                color="black",
                fontsize=9
            )
            if "Все категории" in categories or len(categories) == 1:
                category = "Суммарно"
            else:
                label = sel.artist.get_label()
                category = label.replace(": прогноз", "").strip()
            x_data = sel.target[0]
            idx = int(round(x_data))
            if 0 <= idx < len(month_labels):
                month = month_labels[idx]
                value = int(round(sel.target[1]))
                sel.annotation.set_text(
                    f"Категория: {category}\nМесяц: {month}\nВыручка: {value:,} ₽".replace(',', ' ')
                )
            else:
                sel.annotation.set_visible(False)
                self.canvas2.draw_idle()
            if hasattr(self, '_tooltip_hide_timer_fig2'): #Таймер автоматического скрытия
                self.root.after_cancel(self._tooltip_hide_timer_fig2)
            self._tooltip_hide_timer_fig2 = self.root.after(
                1000, lambda s=sel: self._hide_tooltip_safe_fig2(s)
            )

    def show_seasonality(self): #Переключает интерфейс на вкладку "Сезонность"
        self.current_view = "seasonality"
        self.charts_frame.pack_forget() #Скрываем основные графики и статистику
        self.elasticity_frame.pack_forget()
        self.stats_frame.pack_forget()
        self.seasonality_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5) #Отображаем фрейм сезонности
        if self.df_full is not None and not self.df_full.empty: #Обновляем данные и перерисовываем график
            seasonal_coeffs, _ = self._calculate_seasonality(self.df_full)
            self.seasonal_coeffs = seasonal_coeffs
            self._draw_seasonality_graph()
        else:
            self._draw_seasonality_graph()

    def show_elasticity(self): #Переключает интерфейс на вкладку "Эластичность"
        self.current_view = "elasticity"
        self.charts_frame.pack_forget() #Скрываем другие фреймы
        self.seasonality_frame.pack_forget()
        self.stats_frame.pack_forget()
        self.elasticity_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5) #Отображаем фрейм эластичности
        if self.df_full is not None and not self.df_full.empty: #Перерисовываем график
            self._draw_elasticity_graph()
        else:
            self._draw_elasticity_graph()

    def show_charts(self): #Возвращает интерфейс к основным графикам
        self.current_view = "charts"
        self.seasonality_frame.pack_forget() #Скрываем дополнительные вкладки
        self.elasticity_frame.pack_forget()
        self.charts_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5) #Отображаем основные графики и статистику
        self.stats_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        if self.df_full is not None and not self.df_full.empty: #Обновляем основные графики
            self._draw_animation(self.selected_categories)

    def _draw_elasticity_graph(self): #Отрисовывает график ценовой эластичности и эффективности скидок
        if self.df_full is None or self.df_full.empty: #Отображение сообщения при отсутствии данных
            self.ax_elasticity.clear()
            self.ax_elasticity.text(0.5, 0.5, "Нет данных", ha='center', va='center', fontsize=14, color='gray')
            self.canvas_elasticity.draw()
            return
        if hasattr(self, '_elasticity_tooltip_handler'): #Отключение предыдущего обработчика подсказок
            self.fig_elasticity.canvas.mpl_disconnect(self._elasticity_tooltip_handler)
        self.fig_elasticity.clear() #Очистка и подготовка холста
        self.ax_elasticity = self.fig_elasticity.add_subplot(111)
        df = self.df_full.copy() #Подготовка данных
        df['Скидки'] = pd.to_numeric(df['Скидки'], errors='coerce').fillna(0)
        df['Цена со скидкой'] = df['Цена за шт'] * (1 - df['Скидки'] / 100)
        df['Выручка'] = df['Цена со скидкой'] * df['Количество продаж']
        if "Все категории" in self.selected_categories: #Фильтрация категорий для отображения
            categories = [ #Отображаем только категории с сильной эластичностью
                cat for cat, el in self.price_elasticity.items()
                if abs(el) > 1.0
            ]
            if not categories:
                self.ax_elasticity.clear()
                self.ax_elasticity.text(
                    0.5, 0.5,
                    "Нет категорий с сильной ценовой эластичностью\n(|эластичность| ≤ 1.0 для всех)",
                    ha='center', va='center', fontsize=14, color='gray'
                )
                self.canvas_elasticity.draw()
                return
        else: #Отображаем выбранные категории
            categories = self.selected_categories
        markers = ['o', 's', '^', 'D', 'v'] #Настройка визуальных параметров
        colors = ['#3e7cb4', '#2c6c7e', '#3e55b4', '#1a414c', '#11442e']
        sizes = np.sqrt(df['Выручка']) * 2
        scatter_objects = [] #Список scatter-объектов для анимации
        legend_elements = [] #Элементы легенды
        self._elasticity_tooltip_data = [] #Данные для подсказок
        for idx, cat in enumerate(categories): #Отрисовка scatter-точек по категориям
            cat_data = df[df['Категория'] == cat]
            if cat_data.empty:
                continue
            marker = markers[idx % len(markers)]
            color = colors[idx % len(colors)]
            scatter = self.ax_elasticity.scatter( #Создание scatter-графика
                cat_data['Цена со скидкой'],
                cat_data['Количество продаж'],
                c=cat_data['Скидки'],
                s=sizes[cat_data.index],
                cmap='coolwarm',
                alpha=0.0,
                edgecolors=color,
                linewidth=1.2,
                marker=marker,
                label=cat
            )
            scatter_objects.append(scatter)
            for _, row in cat_data.iterrows(): #Сохранение данных для подсказок
                self._elasticity_tooltip_data.append({
                    'x': row['Цена со скидкой'],
                    'y': row['Количество продаж'],
                    'cat': row['Категория'],
                    'price': row['Цена со скидкой'],
                    'qty': row['Количество продаж'],
                    'discount': row['Скидки'],
                    'revenue': row['Выручка']
                })
            legend_elements.append( #Элементы легенды
                Line2D([0], [0], marker=marker, color='w', label=cat,
                       markerfacecolor=color, markersize=8, linestyle='None')
            )
        self.ax_elasticity.set_xlabel('Цена со скидкой (₽)', fontsize=11) #Настройка осей и заголовка
        self.ax_elasticity.set_ylabel('Спрос (шт)', fontsize=11)
        self.ax_elasticity.set_title('Ценовая эластичность и эффективность скидок', fontsize=13, fontweight='bold')
        self.ax_elasticity.grid(True, linestyle='--', alpha=0.5)
        legend = self.ax_elasticity.legend( #Создание легенды
            handles=legend_elements,
            loc='upper left',
            bbox_to_anchor=(0.0, -0.27),
            fontsize=8,
            ncol=1 if len(legend_elements) < 5 else 2,
            frameon=True
        )
        legend.get_frame().set_alpha(0.9)
        if self.dark_theme: #Настройка цветовой палитры в зависимости от темы
            legend_bg = "#3a3f4b"
            legend_edge = "#555"
            legend_text = "white"
            bg = "#1e1e1e"
            fg = "white"
        else:
            legend_bg = "white"
            legend_edge = "lightgray"
            legend_text = "#2c3e50"
            bg = "white"
            fg = "#2c3e50"
        legend.get_frame().set_facecolor(legend_bg)
        legend.get_frame().set_edgecolor(legend_edge)
        for text in legend.get_texts():
            text.set_color(legend_text)
        if categories: #Добавление цветовой шкалы для скидок
            all_discounts = df[df['Категория'].isin(categories)]['Скидки']
            if not all_discounts.empty:
                from matplotlib.cm import ScalarMappable
                from matplotlib.colors import Normalize
                norm = Normalize(vmin=all_discounts.min(), vmax=all_discounts.max()) #Нормализация значений скидок
                sm = ScalarMappable(cmap='coolwarm', norm=norm)
                sm.set_array([])
                cbar = plt.colorbar(sm, ax=self.ax_elasticity) #Создание цветовой шкалы
                cbar.set_label('Скидка (%)', fontsize=10)
                self._elasticity_colorbar = cbar
                cbar.ax.yaxis.set_tick_params(color=fg) #Настройка цветов шкалы под текущую тему
                cbar.outline.set_edgecolor(fg)
                plt.setp(cbar.ax.get_yticklabels(), color=fg)
                cbar.ax.yaxis.label.set_color(fg)
        self.fig_elasticity.patch.set_facecolor(bg) #Применение цветовой темы к графику
        self.ax_elasticity.set_facecolor(bg)
        self.ax_elasticity.tick_params(colors=fg)
        self.ax_elasticity.xaxis.label.set_color(fg)
        self.ax_elasticity.yaxis.label.set_color(fg)
        self.ax_elasticity.title.set_color(fg)
        for spine in self.ax_elasticity.spines.values():
            spine.set_color(fg)
        self.fig_elasticity.subplots_adjust(bottom=0.37) #Настройка отступов для легенды
        self.anim_elasticity_scatters = scatter_objects
        self.canvas_elasticity.draw()
        self._tooltip_annotation = self.ax_elasticity.annotate( #Настройка подсказок при наведении
            "", xy=(0, 0), xytext=(10, 10),
            textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.3", fc="white", alpha=0.9),
            arrowprops=dict(arrowstyle="->"),
            zorder=1000
        )
        self._tooltip_annotation.set_visible(False)

        def on_hover(event): #Обработчик движения мыши для подсказок
            vis = self._tooltip_annotation.get_visible()
            if event.inaxes != self.ax_elasticity:
                if vis:
                    self._tooltip_annotation.set_visible(False)
                    self.canvas_elasticity.draw_idle()
                return
            found = False
            for point in self._elasticity_tooltip_data:
                dx = event.xdata - point['x']
                dy = event.ydata - point['y']
                dist = np.hypot(dx, dy)
                if dist < 5: #Порог расстояния для активации подсказки
                    xlim = self.ax_elasticity.get_xlim() #Автоматическое позиционирование подсказки, чтобы не выходила за границы графика
                    ylim = self.ax_elasticity.get_ylim()
                    offset_x, offset_y = 10, 10
                    if point['x'] > xlim[1] * 0.7:
                        offset_x = -120
                    elif point['x'] < xlim[0] * 1.2:
                        offset_x = 60
                    if point['y'] > ylim[1] * 0.8:
                        offset_y = -60
                    elif point['y'] < ylim[0] * 1.2:
                        offset_y = 40
                    self._tooltip_annotation.xy = (point['x'], point['y'])
                    self._tooltip_annotation.set_position((offset_x, offset_y))
                    text = (
                        f"Категория: {point['cat']}\n"
                        f"Цена: {point['price']:.0f} ₽\n"
                        f"Спрос: {point['qty']:.0f} шт\n"
                        f"Скидка: {point['discount']:.1f}%\n"
                        f"Выручка: {point['revenue']:.0f} ₽"
                    )
                    self._tooltip_annotation.set_text(text)
                    self._tooltip_annotation.set_visible(True)
                    found = True
                    break
            if vis or found:
                self._tooltip_annotation.set_visible(found)
                self.canvas_elasticity.draw_idle()
        self._elasticity_tooltip_handler = self.fig_elasticity.canvas.mpl_connect( #Подключение обработчика событий мыши
            "motion_notify_event", lambda event: on_hover(event)
        )
        steps = 10 #Параметры анимации
        delay = 60
        final_alpha = 0.7

        def animate_step(step): #Постепенное увеличение прозрачности scatter-точек
            alpha = (step / steps) * final_alpha
            for scatter in self.anim_elasticity_scatters:
                scatter.set_alpha(alpha)
            self.canvas_elasticity.draw_idle()
            if step < steps:
                self.root.after(delay, animate_step, step + 1)
        self.root.after(100, animate_step, 1)
        recommendations = self._generate_elasticity_recommendations() #Отображение рекомендаций по эластичности
        self.elasticity_text_widget.config(state=tk.NORMAL)
        self.elasticity_text_widget.delete(1.0, tk.END)
        self.elasticity_text_widget.insert(tk.END, recommendations)
        self.elasticity_text_widget.config(state=tk.DISABLED)
        stats = [] #Формирование строки статистики для эластичности
        if hasattr(self, 'price_elasticity') and self.price_elasticity: #Расчёт средней эластичности по всем категориям
            el_values = [v for v in self.price_elasticity.values() if np.isfinite(v)]
            if el_values:
                avg_el = np.mean(el_values)
                stats.append(f"Средняя эластичность: {avg_el:.2f}")
        if hasattr(self, 'discount_effect') and self.discount_effect: #Расчёт среднего эффекта от скидок на выручку
            rev_lifts = [v[1] for v in self.discount_effect.values() if np.isfinite(v[1])]
            if rev_lifts:
                avg_rev_lift = np.mean(rev_lifts)
                stats.append(f"Средний эффект скидок: {avg_rev_lift:+.1f}%")
        self.elasticity_stats_label.config(text=" • ".join(stats) if stats else "Нет данных для агрегации")

    def _draw_seasonality_graph(self): #Отрисовывает график сезонности спроса по месяцам
        if self.df_full is None or self.df_full.empty: #Отображение сообщения при отсутствии данных
            self.ax3.clear()
            self.ax3.text(0.5, 0.5, "Нет данных", ha='center', va='center', fontsize=14, color='gray')
            self.canvas3.draw()
            return
        self.fig3.clear() #Очистка и подготовка холста
        self.ax3 = self.fig3.add_subplot(111)
        df = self.df_full.copy() #Подготовка данных
        df['месяц'] = df['Дата'].dt.month
        df['Категория'] = df['Категория'].fillna('Прочее')
        grouped = df.groupby(['Категория', 'месяц'])['Количество продаж'].sum().reset_index()
        month_names = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                       "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
        x = np.arange(12)
        top_categories = list(self.seasonal_coeffs.keys())[:5] #Выбор топ-5 категорий с выраженной сезонностью
        total_by_month = df.groupby('месяц')['Количество продаж'].sum() #Расчёт коэффициентов сезонности для всех категорий
        avg_total = total_by_month.mean() if not total_by_month.empty else 1
        all_coeffs = [total_by_month.get(i + 1, 0) / avg_total for i in range(12)]
        bars_all = self.ax3.bar(x - 0.15, all_coeffs, width=0.3, label='Все категории', #Отрисовка столбцов для всех категорий
                                color='lightgray', alpha=0.7, edgecolor='gray')
        colors = ['#3e7cb4', '#2c6c7e', '#3e55b4', '#1a414c', '#11442e'] #Настройка цветов для категорий
        max_height = 0
        highlights = []
        bars_lines = []
        for idx, cat in enumerate(top_categories): #Отрисовка столбцов для каждой категории с выраженной сезонностью
            cat_data = grouped[grouped['Категория'] == cat]
            if cat_data.empty:
                continue
            avg = cat_data['Количество продаж'].mean()
            coeffs = [cat_data[cat_data['месяц'] == i + 1]['Количество продаж'].sum() / avg
                      if avg > 0 else 1.0 for i in range(12)]
            offset = 0.15 + idx * 0.12
            bars_cat = self.ax3.bar(x + offset, coeffs, width=0.1,
                                    label=cat, color=colors[idx % len(colors)], alpha=0.9)
            bars_lines.extend(bars_cat)
            max_height = max(max_height, max(coeffs))
            for m, c in enumerate(coeffs): #Сбор информации о пиках для отображения в подсказках
                if c > 1.8:
                    highlights.append(f"{month_names[m]} ×{c:.1f} ({cat})")
        self.ax3.set_xticks(x) #Настройка осей и оформления графика
        self.ax3.set_xticklabels(month_names, fontsize=10)
        self.ax3.set_ylabel('Сезонный коэффициент', fontsize=11)
        self.ax3.set_title('Сезонность спроса по месяцам', fontsize=13, fontweight='bold')
        self.ax3.axhline(1.0, color='black', linestyle='--', linewidth=1, alpha=0.6) #Горизонтальная линия на уровне 1.0
        self.ax3.grid(True, axis='y', linestyle='--', alpha=0.5)
        self.ax3.set_ylim(0, max(2.5, max_height * 1.1))
        legend_elements = [] #Создание легенды
        legend_elements.append(Patch(facecolor='lightgray', alpha=0.7, edgecolor='gray', label='Все категории'))
        for idx, cat in enumerate(top_categories):
            color = colors[idx % len(colors)]
            legend_elements.append(Patch(facecolor=color, alpha=0.9, label=cat))
        legend = self.ax3.legend(
            handles=legend_elements,
            loc='upper left',
            bbox_to_anchor=(0.0, -0.27),
            fontsize=8,
            ncol=1 if len(legend_elements) < 5 else 2,
            frameon=True
        )
        legend.get_frame().set_alpha(0.9)
        if self.dark_theme: #Настройка цветовой палитры в зависимости от темы
            legend_bg = "#3a3f4b"
            legend_edge = "#555"
            legend_text = "white"
            bg = "#1e1e1e"
            fg = "white"
        else:
            legend_bg = "white"
            legend_edge = "lightgray"
            legend_text = "#2c3e50"
            bg = "white"
            fg = "#2c3e50"
        legend.get_frame().set_facecolor(legend_bg)
        legend.get_frame().set_edgecolor(legend_edge)
        for text in legend.get_texts():
            text.set_color(legend_text)
        self.fig3.patch.set_facecolor(bg) #Применение цветовой темы к графику
        self.ax3.set_facecolor(bg)
        self.ax3.tick_params(colors=fg)
        self.ax3.xaxis.label.set_color(fg)
        self.ax3.yaxis.label.set_color(fg)
        self.ax3.title.set_color(fg)
        for spine in self.ax3.spines.values():
            spine.set_color(fg)
        self.fig3.subplots_adjust(bottom=0.37) #Настройка отступов для легенды
        self.anim3_bars = [rect for rect in bars_all] + bars_lines
        self.canvas3.draw()
        if hasattr(self, '_seasonality_tooltip_handler'): #Настройка подсказок для графика сезонности
            self.fig3.canvas.mpl_disconnect(self._seasonality_tooltip_handler)
        self._seasonality_tooltip_data = []
        month_names = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                       "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
        for i, bar in enumerate(bars_all): #Данные для подсказок по столбцам "Все категории"
            self._seasonality_tooltip_data.append({
                'x': bar.get_x() + bar.get_width() / 2,
                'y': bar.get_height(),
                'category': 'Все категории',
                'month': month_names[i],
                'coefficient': round(bar.get_height(), 2)
            })
        bar_idx = 0 #Данные для подсказок по категориям с выраженной сезонностью
        for cat in top_categories:
            cat_data = grouped[grouped['Категория'] == cat]
            if cat_data.empty:
                continue
            avg = cat_data['Количество продаж'].mean()
            coeffs = [
                cat_data[cat_data['месяц'] == m + 1]['Количество продаж'].sum() / avg
                if avg > 0 else 1.0 for m in range(12)
            ]
            for m in range(12):
                if bar_idx < len(bars_lines):
                    bar = bars_lines[bar_idx]
                    self._seasonality_tooltip_data.append({
                        'x': bar.get_x() + bar.get_width() / 2,
                        'y': bar.get_height(),
                        'category': cat,
                        'month': month_names[m],
                        'coefficient': round(coeffs[m], 2)
                    })
                    bar_idx += 1
        self._seasonality_tooltip_annot = self.ax3.annotate( #Создание аннотации для подсказок
            "", xy=(0, 0), xytext=(10, 10),
            textcoords="offset points",
            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="black", alpha=0.9),
            arrowprops=dict(arrowstyle="->", color="black"),
            zorder=1000,
            color="black"
        )
        self._seasonality_tooltip_annot.set_visible(False)

        def on_hover_seasonality(event): #Обработчик движения мыши для подсказок сезонности
            vis = self._seasonality_tooltip_annot.get_visible()
            if event.inaxes != self.ax3:
                if vis:
                    self._seasonality_tooltip_annot.set_visible(False)
                    self.canvas3.draw_idle()
                return
            found = False
            for point in self._seasonality_tooltip_data:
                if abs(event.xdata - point['x']) < 0.07:
                    x, y = point['x'], point['y']
                    text = (
                        f"Категория: {point['category']}\n"
                        f"Месяц: {point['month']}\n"
                        f"Коэффициент: {point['coefficient']:.2f}"
                    )
                    xlim = self.ax3.get_xlim() #Автоматическое позиционирование подсказки
                    ylim = self.ax3.get_ylim()
                    x_range = xlim[1] - xlim[0]
                    y_range = ylim[1] - ylim[0]
                    offset_x, offset_y = 10, 10
                    if x > xlim[1] - 0.2 * x_range:
                        offset_x = -10 - self._seasonality_tooltip_annot.get_bbox_patch().get_width() if hasattr(
                            self._seasonality_tooltip_annot.get_bbox_patch(), 'get_width') else -80
                    if y > ylim[1] - 0.15 * y_range:
                        offset_y = -30
                    self._seasonality_tooltip_annot.xy = (x, y)
                    self._seasonality_tooltip_annot.set_text(text)
                    self._seasonality_tooltip_annot.xyann = (offset_x, offset_y)
                    self._seasonality_tooltip_annot.set_visible(True)
                    found = True
                    break
            if vis or found:
                self._seasonality_tooltip_annot.set_visible(found)
                self.canvas3.draw_idle()
        self._seasonality_tooltip_handler = self.fig3.canvas.mpl_connect( #Подключение обработчика событий мыши
            "motion_notify_event", lambda event: on_hover_seasonality(event)
        )
        recommendations = self._generate_seasonality_recommendations() #Отображение рекомендаций по сезонности
        if hasattr(self, 'seasonality_text_widget'):
            self.seasonality_text_widget.config(state=tk.NORMAL)
            self.seasonality_text_widget.delete(1.0, tk.END)
            self.seasonality_text_widget.insert(tk.END, recommendations)
            self.seasonality_text_widget.config(state=tk.DISABLED)
        if hasattr(self, 'seasonal_coeffs') and self.seasonal_coeffs: #Формирование строки статистики для сезонности
            insights = []
            for cat, coeffs in self.seasonal_coeffs.items():
                peak_month, peak_val = max(coeffs.items(), key=lambda x: x[1])
                month_name = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                              "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"][peak_month - 1]
                insights.append(f"{month_name} ×{peak_val:.1f} ({cat})")
            seasonality_summary = " • ".join(insights[:5])
            self.seasonality_stats_label.config(text=f"Сезонность: {seasonality_summary}")
        else:
            self.seasonality_stats_label.config(text="Сезонность: данных недостаточно для анализа")
        steps = 10 #Анимация появления столбцов на графике сезонности
        delay = 60

        def animate_step(step): #Постепенное увеличение прозрачности столбцов
            alpha = step / steps
            for rect in self.anim3_bars:
                rect.set_alpha(alpha)
            self.canvas3.draw_idle()
            if step < steps:
                self.root.after(delay, animate_step, step + 1)
        self.root.after(100, animate_step, 1)

    def _hide_tooltip_safe(self, sel):  #Метод для безопасного скрытия подсказки на первом графике
        try:
            sel.annotation.set_visible(False)
            self.canvas1.draw_idle()
        except Exception:
            pass

    def _hide_tooltip_safe_fig2(self, sel): #Метод для безопасного скрытия подсказки на втором графике
        try:
            sel.annotation.set_visible(False)
            self.canvas2.draw_idle()
        except Exception:
            pass

    def export_to_excel(self): #Экспорт результатов анализа в Excel с 6 листами
        if self.df_full is None:
            messagebox.showwarning("Внимание", "Сначала запустите анализ.")
            return
        try: #Формирование имени файла
            cat_name = "все" if "Все категории" in self.selected_categories else "_".join(self.selected_categories[:3])
            default_name = f"Прогноз_спроса_{sanitize_filename(cat_name)}" + ".xlsx"
            path = filedialog.asksaveasfilename(
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Сохранить прогноз"
            )
            if not path:
                return
            if not path.lower().endswith('.xlsx'):
                path += '.xlsx'
            wb = openpyxl.Workbook() #Создание рабочей книги Excel
            ws1 = wb.active
            ws1.title = "Прогноз на 2026"
            ws1.append(["Месяц", "Прогноз выручки (₽)", "Прогноз спроса (шт)"])
            for cell in ws1[1]: #Стилизация заголовка
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DEEBF7")
                cell.alignment = Alignment(horizontal="center")
            if self.future_monthly is not None and not self.future_monthly.empty: #Заполнение данных прогноза на 2026
                for _, row in self.future_monthly.iterrows():
                    month = str(row['месяц_str']) if 'месяц_str' in row else f"2026-{int(row['month_id'])+1:02d}"
                    rev = int(round(row['прогноз_выручка'])) if pd.notna(row['прогноз_выручка']) else 0
                    qty = int(round(row['прогноз_спроса'])) if pd.notna(row['прогноз_спроса']) else 0
                    ws1.append([month, rev, qty])
            else: #Fallback: агрегация по категориям
                rev_total = np.zeros(12)
                qty_total = np.zeros(12)
                for cat in self.selected_categories:
                    if cat in self.future_by_cat:
                        rev_total += self.future_by_cat[cat]
                for i in range(12):
                    ws1.append([f"2026-{i+1:02d}", int(rev_total[i]), int(rev_total[i] // 100)])
            ws2 = wb.create_sheet("Оценка по месяцам") #Лист 2: Оценка по месяцам
            ws2.append(["Месяц", "Факт: спрос (шт)", "Оценка: спрос (шт)", "Факт: выручка (₽)", "Оценка: выручка (₽)"])
            for cell in ws2[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E2F0D9")
                cell.alignment = Alignment(horizontal="center")
            for _, row in self.monthly_df.iterrows():
                month = str(row['месяц_str'])
                fact_qty = int(row['Количество продаж'])
                pred_qty = int(round(row['прогноз_спроса'])) if 'прогноз_спроса' in row and pd.notna(row['прогноз_спроса']) else fact_qty
                fact_rev = int(fact_qty * row['Цена за шт'])
                pred_rev = int(round(row['оценка_выручки'])) if 'оценка_выручки' in row and pd.notna(row['оценка_выручки']) else fact_rev
                ws2.append([month, fact_qty, pred_qty, fact_rev, pred_rev])
            ws3 = wb.create_sheet("По категориям") #Лист 3: По категориям
            ws3.append(["Категория", "Факт: спрос", "Прогноз: спрос", "Факт: выручка", "Прогноз: выручка"])
            for cell in ws3[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D6EAF8")
                cell.alignment = Alignment(horizontal="center")
            for cat in self.selected_categories:
                if cat == "Все категории":
                    cat_df = self.df_full
                else:
                    cat_df = self.df_full[self.df_full['Категория'] == cat]
                if cat_df.empty:
                    continue
                fact_qty = cat_df['Количество продаж'].sum()
                pred_qty = cat_df['прогноз_спроса'].sum() if 'прогноз_спроса' in cat_df.columns else fact_qty
                fact_rev = (cat_df['Количество продаж'] * cat_df['Цена за шт']).sum()
                pred_rev = cat_df['оценка_выручки'].sum() if 'оценка_выручки' in cat_df.columns else fact_rev
                ws3.append([cat, int(fact_qty), int(pred_qty), int(fact_rev), int(pred_rev)])
            ws4 = wb.create_sheet("Товары") #Лист 4: По товарам
            ws4.append(["Товар", "Категория", "Средняя цена", "Факт: спрос", "Прогноз: спрос", "Факт: выручка", "Прогноз: выручка"])
            for cell in ws4[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.alignment = Alignment(horizontal="center")
            df_export = self.df_full.copy()
            if 'прогноз_спроса' not in df_export.columns:
                df_export['прогноз_спроса'] = df_export['Количество продаж']
            if 'оценка_выручки' not in df_export.columns:
                df_export['оценка_выручки'] = df_export['Количество продаж'] * df_export['Цена за шт']
            item_agg = df_export.groupby('Товар', as_index=False).agg({
                'Категория': 'first',
                'Цена за шт': 'mean',
                'Количество продаж': 'sum',
                'прогноз_спроса': 'sum',
                'оценка_выручки': 'sum'
            })
            for _, row in item_agg.iterrows():
                ws4.append([
                    row['Товар'],
                    row['Категория'],
                    round(row['Цена за шт'], 2),
                    int(row['Количество продаж']),
                    int(round(row['прогноз_спроса'])),
                    int(round(row['Количество продаж'] * row['Цена за шт'])),
                    int(round(row['оценка_выручки']))
                ])
            ws5 = wb.create_sheet("Сезонность") #Лист 5: Сезонность
            ws5.append(["Месяц", "Все категории"] + sorted(self.seasonal_coeffs.keys()) if hasattr(self,
                                                                                                   'seasonal_coeffs') else [])
            for cell in ws5[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E2F0D9")
                cell.alignment = Alignment(horizontal="center")
            month_names = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                           "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
            for i, month in enumerate(range(1, 13)):
                row = [month_names[i]]
                if hasattr(self, 'df_full') and not self.df_full.empty:
                    df = self.df_full
                    df['месяц'] = df['Дата'].dt.month
                    total_by_month = df.groupby('месяц')['Количество продаж'].sum()
                    avg_total = total_by_month.mean()
                    all_coeff = total_by_month.get(month, 0) / avg_total if avg_total else 1.0
                    row.append(round(all_coeff, 2))
                else:
                    row.append(1.0)
                if hasattr(self, 'seasonal_coeffs'):
                    for cat in sorted(self.seasonal_coeffs.keys()):
                        coeff = self.seasonal_coeffs[cat].get(month, 1.0)
                        row.append(round(coeff, 2))
                else:
                    for _ in sorted(self.seasonal_coeffs.keys()):
                        row.append(1.0)
                ws5.append(row)
            for row in ws5.iter_rows(min_row=2, max_row=13, min_col=2, max_col=ws5.max_column): #Условное форматирование: выделение пиков и провалов
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value > 1.5:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE") #Пик сезонности
                        cell.font = Font(bold=True, color="006100")
                    elif isinstance(cell.value, (int, float)) and cell.value < 0.7:
                        cell.fill = PatternFill("solid", fgColor="F8CBAD") #Провал сезонности
                        cell.font = Font(color="9C5700")
            ws5.append([]) #Добавление рекомендаций по сезонности
            ws5.append(["Рекомендации"])
            ws5.merge_cells(start_row=ws5.max_row, start_column=1, end_row=ws5.max_row, end_column=ws5.max_column)
            ws5[ws5.max_row][0].font = Font(bold=True, color="1A5276")
            ws5[ws5.max_row][0].alignment = Alignment(horizontal="left")
            recs = []
            if hasattr(self, 'seasonal_coeffs'):
                for cat, coeffs in self.seasonal_coeffs.items():
                    peaks = [m for m, c in coeffs.items() if c > 1.8]
                    lows = [m for m, c in coeffs.items() if c < 0.6]
                    if peaks:
                        peak_months = ", ".join([month_names[m - 1] for m in sorted(peaks)])
                        recs.append(f"• {cat}: повышенный спрос в {peak_months} → увеличить закупки за 1–2 мес.")
                    if lows:
                        low_months = ", ".join([month_names[m - 1] for m in sorted(lows)])
                        recs.append(f"• {cat}: низкий спрос в {low_months} → избегать закупок, запускать акции.")
            if not recs:
                recs = ["• Чёткой сезонности не обнаружено. Рекомендуется собрать больше данных (≥2 года)."]
            for rec in recs:
                ws5.append([rec])
            ws6 = wb.create_sheet("Эластичность и скидки") #Лист 6: Эластичность и скидки
            ws6.append(
                ["Категория", "Эластичность", "Δ спрос при скидке, %", "Δ выручка при скидке, %", "Рекомендация"])
            for cell in ws6[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D0ECE7")
                cell.alignment = Alignment(horizontal="center")
            all_cats = set()
            if hasattr(self, 'price_elasticity'):
                all_cats.update(self.price_elasticity.keys())
            if hasattr(self, 'discount_effect'):
                all_cats.update(self.discount_effect.keys())
            for cat in sorted(all_cats):
                elasticity = self.price_elasticity.get(cat, 0)
                delta_qty, delta_rev = self.discount_effect.get(cat, (0, 0))
                if delta_rev > 5: #Формирование рекомендаций по скидкам
                    rec = "Увеличить охват скидок"
                elif delta_rev < -3:
                    rec = "Пересмотреть механику скидок"
                else:
                    rec = "Оставить без изменений"
                ws6.append([cat, round(elasticity, 2), round(delta_qty, 1), round(delta_rev, 1), rec])
            for row in ws6.iter_rows(min_row=2, max_row=ws6.max_row, min_col=4, max_col=4): #Условное форматирование для эффекта скидок
                cell = row[0]
                if isinstance(cell.value, (int, float)):
                    if cell.value > 5:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
                        cell.font = Font(color="006100")
                    elif cell.value < -3:
                        cell.fill = PatternFill("solid", fgColor="F8CBAD")
                        cell.font = Font(color="9C5700")
            wb.save(path) #Сохранение файла
            messagebox.showinfo("Успех", f"Прогноз сохранён:\n{os.path.basename(path)}")
        except PermissionError:
            messagebox.showerror("Ошибка доступа", "Нет прав на запись. Попробуйте папку «Документы».")
        except Exception as e:
            err_msg = str(e)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("Ошибка экспорта", msg))

    def export_plots(self): #Экспорт графиков в PNG или PDF с возможностью выбора конкретных графиков
        try: #Создание диалогового окна выбора графиков
            dialog = tk.Toplevel(self.root)
            dialog.title("Выберите графики для сохранения")
            dialog.geometry("300x200")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.update_idletasks()
            x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (dialog.winfo_width() // 2)
            y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (dialog.winfo_height() // 2)
            dialog.geometry(f"+{x}+{y}")
            tk.Label(dialog, text="Выберите графики:", font=("Arial", 11)).pack(pady=10)
            var1 = tk.BooleanVar(value=True)
            var2 = tk.BooleanVar(value=True)
            var3 = tk.BooleanVar(value=True)
            var4 = tk.BooleanVar(value=True)
            cb1 = tk.Checkbutton(dialog, text="Факт и оценка", variable=var1)
            cb2 = tk.Checkbutton(dialog, text="Прогноз 2026", variable=var2)
            cb3 = tk.Checkbutton(dialog, text="Сезонность", variable=var3)
            cb4 = tk.Checkbutton(dialog, text="Эластичность", variable=var4)
            for cb in (cb1, cb2, cb3, cb4):
                cb.pack(anchor="w", padx=30, pady=2)
            selected = []

            def on_ok():
                selected.extend([
                    ("факт_и_оценка", var1.get()),
                    ("прогноз_2026", var2.get()),
                    ("сезонность", var3.get()),
                    ("эластичность", var4.get())
                ])
                dialog.destroy()

            def on_cancel():
                dialog.destroy()
            btn_frame = tk.Frame(dialog)
            btn_frame.pack(pady=10)
            tk.Button(btn_frame, text="OK", command=on_ok, width=10).pack(side=tk.LEFT, padx=5)
            tk.Button(btn_frame, text="Отмена", command=on_cancel, width=10).pack(side=tk.LEFT, padx=5)
            self.root.wait_window(dialog)
            if not any(v for _, v in selected):
                return
            cat_name = "графики" #Формирование имени файла
            if "Все категории" not in self.selected_categories:
                cat_name = "_".join(self.selected_categories[:2])
            path = filedialog.asksaveasfilename(
                initialfile=f"Графики_{sanitize_filename(cat_name)}",
                defaultextension=".png",
                filetypes=[("PNG", "*.png"), ("PDF", "*.pdf")],
                title="Сохранить графики"
            )
            if not path:
                return
            base, ext = os.path.splitext(path)
            saved_files = []
            if var1.get(): #Сохранение выбранных графиков
                p = f"{base}_факт_и_оценка{ext}"
                self.fig1.savefig(p, dpi=150, bbox_inches='tight')
                saved_files.append(os.path.basename(p))
            if var2.get():
                p = f"{base}_прогноз_2026{ext}"
                self.fig2.savefig(p, dpi=150, bbox_inches='tight')
                saved_files.append(os.path.basename(p))
            if var3.get():
                p = f"{base}_сезонность{ext}"
                self.fig3.savefig(p, dpi=150, bbox_inches='tight')
                saved_files.append(os.path.basename(p))
            if var4.get():
                p = f"{base}_эластичность{ext}"
                self.fig_elasticity.savefig(p, dpi=150, bbox_inches='tight')
                saved_files.append(os.path.basename(p))
            messagebox.showinfo("Готово", f"Сохранено:\n" + "\n".join(saved_files))
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить графики:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = SalesAnalyzerApp(root)
    root.mainloop()